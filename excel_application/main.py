def process_excel(
    excel_file,
    output_folder,
    filename,
    allowed_late_evening_staff=None,
    allowed_late_morning_staff=None,
):
    import pandas as pd
    from datetime import datetime, time
    import re
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.styles import Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    import os
    from decimal import Decimal, ROUND_HALF_UP, getcontext

    getcontext().prec = 10  # Độ chính xác
    from datetime import datetime, time, timedelta

    def lam_tron_don_gia(rate: Decimal) -> Decimal:
        so = int(rate)
        tail = so % 1000
        if str(tail).startswith(("9", "8")):
            return Decimal(str(((so + 999) // 1000) * 1000))
        return rate

    def normalize_name(name: str) -> str:
        # Loại bỏ số thứ tự đầu (ví dụ: 4_), gạch dưới -> khoảng trắng, strip, title case
        name = re.sub(r"^\d+_", "", name)  # Xóa số và dấu gạch dưới đầu
        name = name.replace("_", " ").strip()  # Thay _ bằng khoảng trắng
        return name.title()  # Viết hoa chữ cái đầu

    if allowed_late_evening_staff is None:
        allowed_late_evening_staff = {}
    if allowed_late_morning_staff is None:
        allowed_late_morning_staff = {}

    # Chuẩn hóa tên nhân viên từ UI input
    allowed_late_evening_staff = {
        normalize_name(name): shifts
        for name, shifts in allowed_late_evening_staff.items()
    }

    allowed_late_morning_staff = {
        normalize_name(name): shifts
        for name, shifts in allowed_late_morning_staff.items()
    }

    # Ca chính:
    start_time_morning = time(7, 0)
    end_time_morning = time(15, 0)

    start_time_evening = time(15, 0)
    end_time_evening = time(23, 0)

    # Function
    def fine_late_checkin(so_phut_tre):
        if so_phut_tre <= 0:
            return 0
        elif so_phut_tre < 30:
            return so_phut_tre * 5000
        elif so_phut_tre == 30:
            return 200_000
        else:
            return 200_000 + (so_phut_tre - 30) * 5000

    def parse_duration(duration_str):
        if pd.isna(duration_str):
            return 0.0
        duration_str = str(duration_str)

        hours = 0
        minutes = 0
        match_hour = re.search(r"(\d+)\s*giờ", duration_str)
        match_min = re.search(r"(\d+)\s*phút", duration_str)

        if match_hour:
            hours = int(match_hour.group(1))
        if match_min:
            minutes = int(match_min.group(1))

        return hours + minutes / 60

    def safe_parse_time(t):
        if pd.isna(t):
            return None
        if isinstance(t, datetime):
            return t.time()
        try:
            return datetime.strptime(str(t).strip(), "%H:%M").time()
        except:
            return None

    def find_matched_shift(
        checkin_time: time, shift_times: list[time], end_of_shift: time
    ) -> tuple[time, str] | None:
        """
        Tìm ca phù hợp theo quy tắc:
        - Nếu check-in cách ca từ -15p đến +45p → khớp ca đó
        - Nếu check-in > ca cuối và < 23:00 → chọn ca cuối, xem là đi trễ
        """
        now = datetime.combine(datetime.today(), checkin_time)
        sorted_shifts = sorted(shift_times)

        for shift_time in sorted_shifts:
            shift_dt = datetime.combine(datetime.today(), shift_time)
            diff = (now - shift_dt).total_seconds() / 60  # phút

            if -15 <= diff <= 45:
                if diff < 0:
                    return shift_time, "early"
                elif diff == 0:
                    return shift_time, "on_time"
                else:
                    return shift_time, "late"

        # Nếu đi sau ca cuối và vẫn trong giờ làm (trước 23:00)
        last_shift = sorted_shifts[-1]
        if checkin_time > last_shift and checkin_time < end_of_shift:
            return last_shift, "late"

        return None  # Không phù hợp

    ##

    # ----------------------------------------------------------------------------------------------------------------------
    # Đọc file
    xlsx = pd.ExcelFile(excel_file)
    all_results = []

    for sheet_name in xlsx.sheet_names:
        df_raw = pd.read_excel(xlsx, sheet_name=sheet_name, header=None)

        # Xử lý tên nhân viên bỏ số_
        name_key = normalize_name(sheet_name)

        is_late_evening = name_key in allowed_late_evening_staff
        is_late_morning = name_key in allowed_late_morning_staff
        evening_times = allowed_late_evening_staff.get(name_key, [])
        morning_times = allowed_late_morning_staff.get(name_key, [])
        allowed_times = allowed_late_evening_staff.get(name_key, [])

        # === Xử lý từng dòng check-in/check-out ===
        for i in range(3, len(df_raw) - 1, 2):
            checkin_row = df_raw.iloc[i]
            checkout_row = df_raw.iloc[i + 1]

            date = checkin_row[0]
            checkin = safe_parse_time(checkin_row[1])
            checkout = safe_parse_time(checkout_row[1])

            if not checkin or not checkout:
                continue

            duration_str = checkin_row[3]
            salary_str = checkin_row[4]

            total_hours = parse_duration(duration_str)
            try:
                salary_decimal = Decimal(str(salary_str))
            except:
                salary_decimal = Decimal("0")

            try:
                duration_decimal = Decimal(str(total_hours))
            except:
                duration_decimal = Decimal("0")

            if duration_decimal > 0:
                raw_rate = (salary_decimal / duration_decimal).quantize(
                    Decimal("1."), rounding=ROUND_HALF_UP
                )
                hourly_rate = lam_tron_don_gia(raw_rate)
            else:
                hourly_rate = Decimal("0")

            ghi_chu = []
            worked_hours_morning = 0
            worked_hours_evening = 0
            actual_start_morning = actual_end_morning = None
            actual_start_evening = actual_end_evening = None
            soon_evening = False
            penalty = 0
            matched_shift_found = False

            # === Ca sáng bình thường ===
            if (
                checkin < end_time_morning
                and checkin < time(14, 30)
                and not is_late_morning
            ):
                actual_start_morning = max(checkin, start_time_morning)
                actual_end_morning = min(checkout, end_time_morning)

                if actual_end_morning > actual_start_morning:
                    dt_start = datetime.combine(datetime.today(), actual_start_morning)
                    dt_end = datetime.combine(datetime.today(), actual_end_morning)
                    worked_hours_morning = Decimal(
                        (dt_end - dt_start).total_seconds()
                    ) / Decimal("3600")

                    # Phạt đi trễ ca sáng
                    if checkin > start_time_morning:
                        late_minutes = (
                            datetime.combine(datetime.today(), checkin)
                            - datetime.combine(datetime.today(), start_time_morning)
                        ).total_seconds() // 60
                        phat = fine_late_checkin(late_minutes)
                        penalty += phat
                        if late_minutes < 30:
                            ghi_chu.append(
                                f"Đi trễ {late_minutes} phút (-{phat:,} VND, 5K/phút)"
                            )
                        elif late_minutes == 30:
                            ghi_chu.append(f"Đi trễ 30 phút (-200,000 VND)")
                        else:
                            ghi_chu.append(
                                f"Đi trễ {late_minutes} phút (-{phat:,} VND, 200K + 5K/phút sau 30p)"
                            )

                matched_shift_found = True

            # === Ca sáng gãy ===
            if (
                is_late_morning
                and actual_start_morning is None
                and checkin < time(14, 45)
            ):
                match = find_matched_shift(checkin, morning_times, end_time_morning)
                if match:
                    matched_shift, status = match
                    dt_checkin = datetime.combine(datetime.today(), checkin)
                    dt_shift = datetime.combine(datetime.today(), matched_shift)
                    dt_checkout = datetime.combine(datetime.today(), checkout)

                    # Ngăn overlap với ca chiều
                    if checkin >= time(14, 45):
                        # coi như ca chiều, bỏ qua ca sáng
                        pass

                    elif dt_checkout >= dt_shift + timedelta(minutes=30):
                        actual_start_morning = max(checkin, matched_shift)
                        actual_end_morning = min(checkout, end_time_morning)

                        if actual_end_morning > actual_start_morning:
                            dt_start = datetime.combine(
                                datetime.today(), actual_start_morning
                            )
                            dt_end = datetime.combine(
                                datetime.today(), actual_end_morning
                            )
                            worked_hours_morning = Decimal(
                                (dt_end - dt_start).total_seconds()
                            ) / Decimal("3600")

                            if status == "early":
                                early_minutes = int(
                                    (dt_shift - dt_checkin).total_seconds() // 60
                                )
                                ghi_chu.append(
                                    f"Vào sớm ca {matched_shift.strftime('%H:%M')} ({early_minutes} phút), tính từ {matched_shift.strftime('%H:%M')}"
                                )
                            elif status == "late":
                                late_minutes = int(
                                    (dt_checkin - dt_shift).total_seconds() // 60
                                )
                                phat = fine_late_checkin(late_minutes)
                                penalty += phat
                                ghi_chu.append(
                                    f"Đi trễ {late_minutes} phút (-{phat:,} VND)"
                                )

                            ghi_chu.append(
                                f"Ca gãy sáng: {matched_shift.strftime('%H:%M')} → 15:00"
                            )
                        matched_shift_found = True
                    else:
                        # không đủ 30 phút → chỉ cảnh báo
                        ghi_chu.append(
                            f"Check-in {checkin.strftime('%H:%M')} nhưng checkout quá sớm (<30p)"
                        )
                        warning_flag = True

            # === Ca chiều sớm ====
            if (
                time(14, 45) <= checkin < time(15, 0)
                and checkout >= time(15, 30)
                and not is_late_evening
            ):
                ghi_chu.append("Vào sớm ca chiều: chỉ tính công từ 15:00")
                soon_evening = True
                actual_start_evening = start_time_evening
                actual_end_evening = min(checkout, end_time_evening)

                if actual_end_evening > actual_start_evening:
                    dt_start = datetime.combine(datetime.today(), actual_start_evening)
                    dt_end = datetime.combine(datetime.today(), actual_end_evening)
                    worked_hours_evening = Decimal(
                        (dt_end - dt_start).total_seconds()
                    ) / Decimal("3600")
                    ghi_chu.append("Ca chiều: vào từ 15:00")
                matched_shift_found = True

            # === Ca chiều Thường ===
            # 💡 Nhân viên thường: 2. ca chiều đúng giờ
            if (
                start_time_evening <= checkin < end_time_evening
                and soon_evening == False
                and not is_late_evening
            ):
                actual_start_evening = max(checkin, start_time_evening)
                actual_end_evening = min(checkout, end_time_evening)

                if actual_end_evening > actual_start_evening:

                    dt_start = datetime.combine(datetime.today(), actual_start_evening)
                    dt_end = datetime.combine(datetime.today(), actual_end_evening)
                    worked_hours_evening = Decimal(
                        (dt_end - dt_start).total_seconds()
                    ) / Decimal("3600")

                    if checkin > start_time_evening:
                        late_minutes = (
                            datetime.combine(datetime.today(), checkin)
                            - datetime.combine(datetime.today(), start_time_evening)
                        ).total_seconds() // 60
                        phat = fine_late_checkin(late_minutes)
                        penalty += phat

                        if late_minutes < 30:
                            ghi_chu.append(
                                f"Đi trễ {late_minutes} phút (-{phat:,} VND, 5K/phút)"
                            )
                        elif late_minutes == 30:
                            ghi_chu.append("Đi trễ 30 phút (-200,000 VND)")
                        else:
                            ghi_chu.append(
                                f"Đi trễ {late_minutes} phút (-{phat:,} VND, 200K + 5K/phút sau 30p)"
                            )

                    ghi_chu.append("Ca chiều thường: vào từ 15:00")

                matched_shift_found = True

            # === Ca chiều gãy ===
            if is_late_evening and actual_start_evening is None and not soon_evening:
                # chieu thường của nhân viên ca gãy
                if (
                    checkin < min(allowed_times)
                    and checkin
                    >= (
                        datetime.combine(datetime.today(), min(allowed_times))
                        - timedelta(minutes=30)
                    ).time()
                ):
                    matched_shift = min(allowed_times)
                    actual_start_evening = matched_shift
                    actual_end_evening = min(checkout, end_time_evening)
                    dt_checkin = datetime.combine(datetime.today(), checkin)
                    early_minutes = int(
                        (
                            datetime.combine(datetime.today(), matched_shift)
                            - dt_checkin
                        ).total_seconds()
                        // 60
                    )
                    worked_hours_evening = Decimal(
                        (
                            datetime.combine(datetime.today(), actual_end_evening)
                            - datetime.combine(datetime.today(), actual_start_evening)
                        ).total_seconds()
                    ) / Decimal("3600")
                    ghi_chu.append(
                        f"Vào sớm ca {matched_shift.strftime('%H:%M')} ({early_minutes} phút), tính từ {matched_shift.strftime('%H:%M')}"
                    )
                    ghi_chu.append(f"Ca gãy: {matched_shift.strftime('%H:%M')} → 23:00")
                    matched_shift_found = True
                # ca gãy

                elif match := find_matched_shift(
                    checkin, evening_times, end_time_evening
                ):
                    matched_shift, status = match
                    dt_checkin = datetime.combine(datetime.today(), checkin)
                    dt_shift = datetime.combine(datetime.today(), matched_shift)
                    dt_checkout = datetime.combine(datetime.today(), checkout)

                    if dt_checkout >= dt_shift + timedelta(minutes=30):
                        actual_start_evening = max(checkin, matched_shift)
                        actual_end_evening = min(checkout, end_time_evening)

                        if actual_end_evening > actual_start_evening:
                            dt_start = datetime.combine(
                                datetime.today(), actual_start_evening
                            )
                            dt_end = datetime.combine(
                                datetime.today(), actual_end_evening
                            )
                            worked_hours_evening = Decimal(
                                (dt_end - dt_start).total_seconds()
                            ) / Decimal("3600")

                            if status == "early":
                                early_minutes = int(
                                    (dt_shift - dt_checkin).total_seconds() // 60
                                )
                                ghi_chu.append(
                                    f"Vào sớm ca {matched_shift.strftime('%H:%M')} ({early_minutes} phút), tính từ {matched_shift.strftime('%H:%M')}"
                                )
                            elif status == "late":
                                late_minutes = int(
                                    (dt_checkin - dt_shift).total_seconds() // 60
                                )
                                phat = fine_late_checkin(late_minutes)
                                penalty += phat
                                if late_minutes < 30:
                                    ghi_chu.append(
                                        f"Đi trễ {late_minutes} phút (-{phat:,} VND, 5K/phút)"
                                    )
                                elif late_minutes == 30:
                                    ghi_chu.append("Đi trễ 30 phút (-200,000 VND)")
                                else:
                                    ghi_chu.append(
                                        f"Đi trễ {late_minutes} phút (-{phat:,} VND, 200K + 5K/phút sau 30p)"
                                    )

                            ghi_chu.append(
                                f"Ca gãy: {matched_shift.strftime('%H:%M')} → 23:00"
                            )
                    matched_shift_found = True
            # === Nếu không match ca nào thì cảnh báo ===
            if not matched_shift_found:
                ghi_chu.append(
                    "Đề xuất tính tay vì quá ca gãy trước đó hoặc chấm công quá sớm so với ca gãy tiếp theo (nếu có)"
                )
                warning_flag = True

            worked_hours = worked_hours_morning + worked_hours_evening
            salary = (Decimal(str(worked_hours)) * hourly_rate).quantize(
                Decimal("1"), rounding=ROUND_HALF_UP
            )
            salary_after_penalty = max(Decimal(0), salary - Decimal(penalty))
            note = "; ".join(ghi_chu)

            all_results.append(
                {
                    "Nhân Viên": sheet_name,
                    "Ngày": date,
                    "Chấm Công": checkin,
                    "Kết Thúc Chấm Công": checkout,
                    "Vào Ca Sáng": actual_start_morning,
                    "Ra Ca Sáng": actual_end_morning,
                    "Giờ Làm Ca Sáng": float(
                        Decimal(worked_hours_morning).quantize(Decimal("0.01"))
                    ),
                    "Vào Ca Chiều": actual_start_evening,
                    "Ra Ca Chiều": actual_end_evening,
                    "Giờ Làm Ca Chiều": float(
                        Decimal(worked_hours_evening).quantize(Decimal("0.01"))
                    ),
                    "Tổng giờ công": float(
                        Decimal(worked_hours).quantize(Decimal("0.01"))
                    ),
                    "Tiền Phạt trễ": penalty,
                    "Tiền Công (chưa phạt)": int(salary),
                    "Tiền Công (sau phạt)": int(salary_after_penalty),
                    "Đơn Giá (VND/giờ)": hourly_rate,
                    "Ghi Chú": note,
                }
            )

    result_df = pd.DataFrame(all_results)

    # === Tạo workbook mới ===
    wb = Workbook()
    wb.remove(wb.active)  # Xóa sheet mặc định

    # === 1. Tạo sheet riêng cho từng nhân viên ===
    # Ghi từng sheet nhân viên
    for name, group in result_df.groupby("Nhân Viên"):
        ws = wb.create_sheet(title=name)
        ws.sheet_view.zoomScale = 130
        rows = list(dataframe_to_rows(group, index=False, header=True))

        # Xác định vị trí cột "Ghi Chú"
        header = rows[0]
        try:
            note_idx = header.index("Ghi Chú")
        except ValueError:
            note_idx = -1  # Không có cột ghi chú

        for r_idx, row in enumerate(rows, 1):
            highlight = False
            warning_flag = False  # 👈 thêm flag này

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # Nếu là dòng dữ liệu (không phải tiêu đề)
                if r_idx > 1:
                    if isinstance(value, (int, float, Decimal)):
                        if header[c_idx - 1] in [
                            "Giờ Làm Ca Sáng",
                            "Giờ Làm Ca Chiều",
                            "Tổng giờ công",
                        ]:
                            cell.number_format = "#,##0.00"
                        else:
                            cell.number_format = "#,##0"

                if r_idx > 1 and r_idx % 2 == 0:
                    for c in range(1, len(row) + 1):
                        ws.cell(row=r_idx, column=c).fill = PatternFill(
                            "solid", fgColor="F2F2F2"
                        )

                if r_idx == 1:
                    # Header
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="4F81BD")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    # Kiểm tra ghi chú
                    if note_idx != -1:
                        note_value = row[note_idx]
                        if note_value:
                            note_str = str(note_value)
                            if "Đi trễ" in note_str or "Về sớm" in note_str:
                                highlight = True
                            if "Đề xuất tính tay" in note_str:  # 👈 thêm dòng này
                                warning_flag = True

            # Nếu cần tô đỏ nguyên dòng
            if highlight:
                for c in range(1, len(row) + 1):
                    ws.cell(row=r_idx, column=c).font = Font(color="FF0000")

            # Nếu cảnh báo → tô cam nguyên hàng
            if warning_flag:
                for c in range(1, len(row) + 1):
                    ws.cell(row=r_idx, column=c).fill = PatternFill(
                        "solid", fgColor="FFC000"
                    )  # màu cam

        # Auto-fit cột
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    # === 2. Tạo sheet tổng hợp ===
    summary = (
        result_df.groupby("Nhân Viên")
        .agg(
            {
                "Tổng giờ công": "sum",
                "Tiền Công (chưa phạt)": "sum",
                "Tiền Phạt trễ": "sum",
                "Tiền Công (sau phạt)": "sum",
            }
        )
        .reset_index()
    )

    summary.columns = ["Nhân Viên", "Tổng Giờ", "Tổng Công", "Tổng Phạt", "Thực Nhận"]
    ws_sum = wb.create_sheet(title="Tổng Hợp")
    ws_sum.sheet_view.zoomScale = 130

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for r_idx, row in enumerate(
        dataframe_to_rows(summary, index=False, header=True), 1
    ):
        for c_idx, value in enumerate(row, 1):
            cell = ws_sum.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="9BBB59")
            else:
                if isinstance(value, (int, float, Decimal)):
                    if summary.columns[c_idx - 1] == "Tổng Giờ":
                        cell.number_format = "#,##0.00"
                    else:
                        cell.number_format = "#,##0"
                if r_idx % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F9F9F9")

    # Auto-fit Tổng Hợp
    for col in ws_sum.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws_sum.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    # === 3. Ghi file ===

    output_path = os.path.join(output_folder, filename)
    wb.save(output_path)
