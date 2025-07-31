import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import os
from datetime import datetime
import re
import json
from email.message import EmailMessage
import base64
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import pickle
import openpyxl
from openpyxl import load_workbook, Workbook

# ================== BIẾN TOÀN CỤC ==================
input_file = ""
allowed_late_evening_staff = {}
allowed_late_morning_staff = {}
output_folder = ""

input_file_tab2 = ""
output_folder_tab2 = ""

allowed_late_morning_staff_T12 = {}
allowed_late_evening_staff_T12 = {}

CA_GAY_FILE_T3 = "ca_gay_T3.json"
CA_GAY_FILE_T12 = "ca_gay_T12.json"

# ============ HÀM GỬI MAIL BẰNG GMAIL API ============
SCOPES = ["https://www.googleapis.com/auth/gmail.send"]


def extract_single_sheet(input_path, sheet_name, output_path):
    wb = load_workbook(input_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Không tìm thấy sheet '{sheet_name}' trong file.")

    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = sheet_name
    source_ws = wb[sheet_name]
    for row in source_ws.iter_rows(values_only=True):
        new_ws.append(row)
    new_wb.save(output_path)


def get_gmail_service():
    creds = None
    if os.path.exists("token.pickle"):
        with open("token.pickle", "rb") as token:
            creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file("credentials.json", SCOPES)
            creds = flow.run_local_server(port=0)
        with open("token.pickle", "wb") as token:
            pickle.dump(creds, token)
    service = build("gmail", "v1", credentials=creds)
    return service


def send_email_with_attachment(to, subject, body, attachment_path):
    service = get_gmail_service()
    message = EmailMessage()
    message.set_content(body)
    message["To"] = to
    message["From"] = "me"
    message["Subject"] = subject

    with open(attachment_path, "rb") as f:
        file_data = f.read()
        file_name = os.path.basename(attachment_path)
        message.add_attachment(
            file_data,
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=file_name,
        )

    encoded_message = base64.urlsafe_b64encode(message.as_bytes()).decode()

    create_message = {"raw": encoded_message}

    try:
        send_message = (
            service.users().messages().send(userId="me", body=create_message).execute()
        )
        return send_message
    except HttpError as error:
        print(f"An error occurred: {error}")
        raise error


def create_tab_gui_mail(notebook):

    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
    import os
    import openpyxl

    # Màu sắc và style
    bg_color = "#f4f8fb"  # nền dịu
    border_color = "#b3c6e0"
    group_bg = "#eaf1fb"
    label_fg = "#1a237e"
    entry_bg = "#ffffff"
    btn_color = "#1976D2"
    btn_hover = "#1565c0"
    btn_fg = "#fff"
    border_radius = 10

    tab_mail = tk.Frame(notebook, bg=bg_color, bd=0, highlightthickness=0)
    notebook.add(tab_mail, text="📤 Gửi Mail")

    state = {"excel_path": "", "sheets": [], "selected_sheet": ""}

    font_main = ("Segoe UI", 12)
    font_label = ("Segoe UI", 12, "bold")
    font_entry = ("Segoe UI", 12)
    font_btn = ("Segoe UI", 11, "bold")
    font_title = ("Segoe UI", 16, "bold")

    # Frame tổng
    frame = tk.Frame(tab_mail, bg=bg_color, bd=0, highlightthickness=0)
    frame.pack(padx=30, pady=30, fill="both", expand=True)
    frame.columnconfigure(1, weight=1)

    # Tiêu đề
    lbl_title = tk.Label(
        frame,
        text="📤 Gửi file Excel qua Email",
        font=font_title,
        bg=bg_color,
        fg="#0d47a1",
    )
    lbl_title.grid(row=0, column=0, sticky="w", pady=(0, 18), padx=(0, 0))
    lbl_note = tk.Label(
        frame,
        text="(Gửi email kèm sheet nhân viên có khiếu nại)",
        font=("Segoe UI", 11, "italic"),
        bg=bg_color,
        fg="#b71c1c",
    )
    lbl_note.grid(row=0, column=1, sticky="w", pady=(0, 18), padx=(12, 0))

    # Email người nhận
    tk.Label(
        frame, text="📧 Email người nhận:", font=font_label, bg=bg_color, fg=label_fg
    ).grid(row=1, column=0, sticky="w", pady=8, padx=(0, 8))
    entry_email = tk.Entry(
        frame,
        font=font_entry,
        bg=entry_bg,
        relief="solid",
        bd=1,
        highlightthickness=1,
        highlightbackground=border_color,
    )
    entry_email.grid(
        row=1, column=1, columnspan=2, sticky="we", padx=5, pady=5, ipady=4
    )

    # Group chọn file
    group_file = tk.LabelFrame(
        frame,
        text="📁 Chọn file Excel",
        font=font_label,
        bg=group_bg,
        fg=label_fg,
        bd=0,
        relief="flat",
        labelanchor="nw",
    )
    group_file.grid(
        row=2, column=0, columnspan=3, sticky="we", pady=10, padx=0, ipadx=6, ipady=6
    )
    group_file.columnconfigure(1, weight=1)

    tk.Label(group_file, text="File:", font=font_main, bg=group_bg, fg="#222").grid(
        row=0, column=0, sticky="w", padx=(0, 5)
    )
    lbl_file = tk.Label(
        group_file, text="❌ Chưa chọn", fg="#ff1744", bg=group_bg, font=font_main
    )
    lbl_file.grid(row=0, column=1, sticky="w", padx=5)

    def chon_file_excel():
        file_path = filedialog.askopenfilename(
            title="Chọn file Excel", filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if file_path:
            state["excel_path"] = file_path
            lbl_file.config(text=f"✅ {os.path.basename(file_path)}", fg="#388e3c")
            wb = openpyxl.load_workbook(file_path, data_only=True)
            state["sheets"] = wb.sheetnames
            combo_sheet["values"] = state["sheets"]
            combo_sheet.set(state["sheets"][0])
            state["selected_sheet"] = state["sheets"][0]

    btn_chon = tk.Button(
        group_file,
        text="Chọn file",
        font=font_btn,
        bg=btn_color,
        fg=btn_fg,
        activebackground=btn_hover,
        activeforeground=btn_fg,
        relief="raised",
        bd=2,
        cursor="hand2",
        command=chon_file_excel,
    )
    btn_chon.grid(row=0, column=2, padx=8, ipadx=8, ipady=2)

    # Sheet cần gửi
    tk.Label(
        frame, text="🧾 Tên sheet cần gửi:", font=font_label, bg=bg_color, fg="#222"
    ).grid(row=3, column=0, sticky="w", pady=8, padx=(0, 8))
    combo_sheet = ttk.Combobox(frame, state="readonly", font=font_entry)
    combo_sheet.grid(
        row=3, column=1, columnspan=2, sticky="we", padx=5, pady=5, ipady=2
    )
    combo_sheet.bind(
        "<<ComboboxSelected>>",
        lambda e: state.update({"selected_sheet": combo_sheet.get()}),
    )

    # Tiêu đề email
    tk.Label(
        frame, text="📌 Tiêu đề email:", font=font_label, bg=bg_color, fg="#222"
    ).grid(row=4, column=0, sticky="w", pady=8, padx=(0, 8))
    entry_subject = tk.Entry(
        frame,
        font=font_entry,
        bg=entry_bg,
        relief="solid",
        bd=1,
        highlightthickness=1,
        highlightbackground=border_color,
    )
    entry_subject.grid(
        row=4, column=1, columnspan=2, sticky="we", padx=5, pady=5, ipady=4
    )

    # Nội dung email
    tk.Label(
        frame, text="📝 Nội dung email:", font=font_label, bg=bg_color, fg="#222"
    ).grid(row=5, column=0, sticky="nw", pady=8, padx=(0, 8))
    text_body = tk.Text(
        frame,
        height=7,
        font=font_entry,
        bg=entry_bg,
        relief="solid",
        bd=1,
        highlightthickness=1,
        highlightbackground=border_color,
        wrap="word",
    )
    text_body.grid(row=5, column=1, columnspan=2, sticky="we", padx=5, pady=5)

    # Label trạng thái gửi mail
    status_var = tk.StringVar()
    status_label = tk.Label(
        frame, textvariable=status_var, font=("Segoe UI", 11), fg="#1976d2", bg=bg_color
    )
    status_label.grid(row=7, column=0, columnspan=3, pady=(4, 0), sticky="w")

    def set_status(text, color="#1976d2"):
        status_var.set(text)
        status_label.config(fg=color)

    import threading

    def gui_mail():
        import os

        email = entry_email.get().strip()
        sheet_name = state.get("selected_sheet")
        file_path = state.get("excel_path")
        subject = entry_subject.get().strip()
        body = text_body.get("1.0", tk.END).strip()

        if not email or not file_path or not sheet_name or not subject or not body:
            set_status(
                "⚠️ Vui lòng điền đủ email, tiêu đề, nội dung và chọn file.",
                color="#b71c1c",
            )
            return

        if not email or not file_path or not sheet_name:
            set_status("⚠️ Vui lòng chọn đầy đủ file, sheet và email.", color="#b71c1c")
            return

        def send_mail_task():
            try:
                window.after(
                    0,
                    lambda: set_status(
                        "⏳ Đang gửi mail, vui lòng đợi...", color="#1976d2"
                    ),
                )

                file_tach = f"sheet_{sheet_name}.xlsx"
                wb_goc = openpyxl.load_workbook(file_path, data_only=True)
                sheet = wb_goc[sheet_name]
                wb_moi = openpyxl.Workbook()
                sheet_moi = wb_moi.active
                sheet_moi.title = sheet.title

                for row in sheet.iter_rows(values_only=False):
                    for cell in row:
                        sheet_moi[cell.coordinate].value = cell.value

                wb_moi.save(file_tach)

                send_email_with_attachment(
                    to=email, subject=subject, body=body, attachment_path=file_tach
                )

                def show_success():
                    set_status(f"✅ Đã gửi {file_tach} đến {email}", color="#388e3c")

                window.after(0, show_success)
            except Exception as e:

                def show_error():
                    set_status(f"❌ Gửi mail thất bại: {e}", color="#b71c1c")

                window.after(0, show_error)
            finally:
                # Xóa file tạm sau khi gửi mail (dù thành công hay thất bại)
                try:
                    if os.path.exists(file_tach):
                        os.remove(file_tach)
                except Exception:
                    pass

        threading.Thread(target=send_mail_task, daemon=True).start()

    btn_gui = tk.Button(
        frame,
        text="📤 Gửi Mail",
        font=font_btn,
        bg=btn_color,
        fg=btn_fg,
        activebackground=btn_hover,
        activeforeground=btn_fg,
        relief="raised",
        bd=3,
        cursor="hand2",
        command=gui_mail,
    )
    btn_gui.grid(row=6, column=0, columnspan=3, pady=28, ipadx=18, ipady=6, sticky="we")


def luu_du_lieu_ca_gay_T12():
    data = {
        "sang": {
            k: [t.strftime("%H:%M") for t in v]
            for k, v in allowed_late_morning_staff_T12.items()
        },
        "chieu": {
            k: [t.strftime("%H:%M") for t in v]
            for k, v in allowed_late_evening_staff_T12.items()
        },
    }
    with open("ca_gay_T12.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def tai_du_lieu_ca_gay_T12():
    if os.path.exists(CA_GAY_FILE_T12):
        with open(CA_GAY_FILE_T12, "r", encoding="utf-8") as f:
            du_lieu = json.load(f)
            allowed_late_morning_staff_T12.clear()
            allowed_late_evening_staff_T12.clear()
            allowed_late_morning_staff_T12.update(du_lieu.get("morning", {}))
            allowed_late_evening_staff_T12.update(du_lieu.get("evening", {}))


def xoa_cau_hinh_T12(staff_dict, entry_ten, entry_gio, tree):
    ten = entry_ten.get().strip()
    if ten in staff_dict:
        del staff_dict[ten]
        luu_du_lieu_ca_gay_T12()
        cap_nhat_treeview_T12(tree, staff_dict)
        entry_ten.delete(0, tk.END)
        entry_gio.delete(0, tk.END)


def cap_nhat_treeview_T12(tree, staff_dict):
    tree.delete(*tree.get_children())
    for ten, gio_list in staff_dict.items():
        tree.insert("", "end", values=(ten, ", ".join(gio_list)))


def them_cau_hinh_T12(staff_dict, entry_ten, entry_gio, tree):
    ten = entry_ten.get().strip()
    gio = entry_gio.get().strip()
    if not ten or not gio:
        messagebox.showwarning("Thiếu thông tin", "Vui lòng nhập đầy đủ tên và giờ.")
        return
    gio_list = [g.strip() for g in gio.split(",") if g.strip()]
    staff_dict[ten] = gio_list
    luu_du_lieu_ca_gay_T12()
    cap_nhat_treeview_T12(tree, staff_dict)
    entry_ten.delete(0, tk.END)
    entry_gio.delete(0, tk.END)


def tai_du_lieu_ca_gay(file_path):
    if os.path.exists(file_path):
        with open(file_path, "r") as f:
            return json.load(f)
    return {"morning": {}, "evening": {}}


def luu_du_lieu_ca_gay():
    data = {
        "sang": {
            k: [t.strftime("%H:%M") for t in v]
            for k, v in allowed_late_morning_staff.items()
        },
        "chieu": {
            k: [t.strftime("%H:%M") for t in v]
            for k, v in allowed_late_evening_staff.items()
        },
    }
    with open("ca_gay_T3.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def nap_du_lieu_ca_gay():
    global allowed_late_morning_staff, allowed_late_evening_staff
    try:
        with open("ca_gay_T3.json", "r", encoding="utf-8") as f:
            data = json.load(f)
        allowed_late_morning_staff = {
            k: [datetime.strptime(t, "%H:%M").time() for t in v]
            for k, v in data.get("sang", {}).items()
        }
        allowed_late_evening_staff = {
            k: [datetime.strptime(t, "%H:%M").time() for t in v]
            for k, v in data.get("chieu", {}).items()
        }
    except FileNotFoundError:
        allowed_late_morning_staff = {}
        allowed_late_evening_staff = {}


def nap_du_lieu_ca_gay_T12():
    global allowed_late_morning_staff_T12, allowed_late_evening_staff_T12
    try:
        with open("ca_gay_T12.json", "r", encoding="utf-8") as f:
            data = json.load(f)
        allowed_late_morning_staff_T12 = {
            k: [datetime.strptime(t, "%H:%M").time() for t in v]
            for k, v in data.get("sang", {}).items()
        }
        allowed_late_evening_staff_T12 = {
            k: [datetime.strptime(t, "%H:%M").time() for t in v]
            for k, v in data.get("chieu", {}).items()
        }
    except FileNotFoundError:
        allowed_late_morning_staff_T12 = {}
        allowed_late_evening_staff_T12 = {}


def thoat_chuong_trinh():
    luu_du_lieu_ca_gay()
    luu_du_lieu_ca_gay_T12()
    window.quit()


def normalize_name(name: str) -> str:
    name = re.sub(r"^\d+_", "", name)
    name = name.replace("_", " ").strip()
    return name.title()


def cap_nhat_danh_sach(listbox, data_dict):
    listbox.delete(0, tk.END)
    for ten, times in data_dict.items():
        line = f"{ten}: {', '.join(t.strftime('%H:%M') for t in times)}"
        listbox.insert(tk.END, line)


def convert_data_to_str(data_dict):
    return {
        ten: [t.strftime("%H:%M") for t in times] for ten, times in data_dict.items()
    }


# Xoá cấu hình
def xoa_cau_hinh(listbox, data_dict, title, kv):
    selected = listbox.curselection()
    if not selected:
        messagebox.showwarning("Chưa chọn", "Vui lòng chọn dòng để xoá.")
        return
    line = listbox.get(selected[0])
    ten_nv = line.split(":")[0].strip()
    if ten_nv in data_dict:
        del data_dict[ten_nv]
        cap_nhat_danh_sach(listbox, data_dict)
        messagebox.showinfo("Đã xóa", f"Đã xóa cấu hình {title} của {ten_nv}")

        # Xác định file tương ứng
        if kv == "T3":
            file_path = "ca_gay_T3.json"
        elif kv == "T12":
            file_path = "ca_gay_T12.json"
        else:
            return

        # Đọc file JSON hiện tại (nếu có)
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except FileNotFoundError:
            data = {}

        # Nếu data_dict rỗng sau khi xóa, thì xóa luôn key trong JSON
        if not data_dict:
            if title in data:
                del data[title]
        else:
            data[title] = convert_data_to_str(
                data_dict
            )  # 👈 xử lý datetime trước khi lưu

        # Ghi lại file JSON
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)


# Lưu cấu hình ca
def luu_cau_hinh(
    raw_name, gio_str, data_dict, listbox, title="Chiều", time_format="HH:MM"
):
    ten_nv = normalize_name(raw_name)
    if not ten_nv.strip() or not gio_str.strip():
        messagebox.showwarning("Thiếu thông tin", "Vui lòng nhập tên và giờ.")
        return

    try:
        gio_list = [
            datetime.strptime(g.strip(), "%H:%M").time() for g in gio_str.split(",")
        ]
    except ValueError:
        messagebox.showerror(
            "Lỗi", f"Giờ sai định dạng. Dùng {time_format}, ví dụ: 07:00,08:00"
        )
        return

    data_dict[ten_nv] = sorted(gio_list)
    cap_nhat_danh_sach(listbox, data_dict)

    # ✅ Lưu xuống file sau khi cập nhật
    try:
        luu_du_lieu_ca_gay()
    except Exception as e:
        messagebox.showerror("Lỗi khi lưu", f"Không thể lưu file cấu hình:\n{e}")
        return

    messagebox.showinfo("Đã lưu", f"✅ Đã lưu ca gãy {title} cho: {ten_nv}")


# ================== LẤY TÊN NHÂN VIÊN TỪ EXCEL ==================
def get_staff_names_from_excel(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        # Loại bỏ các sheet mặc định nếu có
        ignore_sheets = {"Sheet", "Sheet1", "Sheet2", "Sheet3"}
        names = [s for s in wb.sheetnames if s not in ignore_sheets]
        return sorted(set(names))
    except Exception:
        return []


# ================== XỬ LÝ FILE ==================
def chon_file_excel():
    global input_file
    file_path = filedialog.askopenfilename(
        title="Chọn file Excel", filetypes=[("Excel Files", "*.xlsx *.xls")]
    )
    if file_path:
        input_file = file_path
        label_file.config(text=f"✅ {os.path.basename(file_path)}")
        btn_chon_thu_muc.config(state="normal")
        entry_ten_file.config(state="normal")
        # Cập nhật danh sách tên nhân viên cho cả Combobox sáng và chiều
        staff_names = get_staff_names_from_excel(file_path)
        combobox_ten_nv_sang_tab1["values"] = staff_names
        combobox_ten_nv_chieu_tab1["values"] = staff_names
        if staff_names:
            combobox_ten_nv_sang_tab1.set(staff_names[0])
            combobox_ten_nv_chieu_tab1.set(staff_names[0])


def chon_file_excel_T12():
    global input_file_tab2
    file_path_2 = filedialog.askopenfilename(
        title="Chọn file Excel", filetypes=[("Excel Files", "*.xlsx *.xls")]
    )
    if file_path_2:
        input_file_tab2 = file_path_2
        label_file2.config(text=f"✅ {os.path.basename(file_path_2)}")
        btn_chon_thu_muc2.config(state="normal")
        entry_ten_file2.config(state="normal")
        # Cập nhật danh sách tên nhân viên cho cả Combobox sáng và chiều tab T12
        staff_names = get_staff_names_from_excel(file_path_2)
        combobox_ten_nv_sang_tab2["values"] = staff_names
        combobox_ten_nv_chieu_tab2["values"] = staff_names
        if staff_names:
            combobox_ten_nv_sang_tab2.set(staff_names[0])
            combobox_ten_nv_chieu_tab2.set(staff_names[0])


def chon_thu_muc():
    global output_folder
    folder = filedialog.askdirectory(title="Chọn thư mục xuất")
    if not folder:
        return
    output_folder = folder
    ten_file = entry_ten_file.get().strip()
    if not ten_file:
        messagebox.showerror("Thiếu tên file", "Vui lòng nhập tên file.")
        return
    if not ten_file.endswith(".xlsx"):
        ten_file += ".xlsx"
    output_path = os.path.join(output_folder, ten_file)
    if os.path.exists(output_path):
        messagebox.showwarning("Trùng tên", "⚠️ Tên file đã tồn tại!")
        return
    label_folder.config(text=f"📁 Nơi lưu: {output_path}")
    status_var_export_t3.set("⏳ Đang xuất file, vui lòng đợi...")
    status_label_export_t3.config(fg="#1976d2")
    window.update_idletasks()
    xu_ly_file_excel(input_file, output_folder, ten_file)


def chon_thu_muc_tab2():
    global output_folder_tab2
    folder = filedialog.askdirectory(title="Chọn thư mục xuất")
    if not folder:
        return
    output_folder_tab2 = folder
    ten_file = entry_ten_file2.get().strip()
    if not ten_file:
        messagebox.showerror("Thiếu tên file", "Vui lòng nhập tên file.")
        return
    if not ten_file.endswith(".xlsx"):
        ten_file += ".xlsx"
    output_path = os.path.join(output_folder_tab2, ten_file)
    if os.path.exists(output_path):
        messagebox.showwarning("Trùng tên", "⚠️ Tên file đã tồn tại!")
        return
    label_folder2.config(text=f"📁 Nơi lưu: {output_path}")
    status_var_export_t12.set("⏳ Đang xuất file, vui lòng đợi...")
    status_label_export_t12.config(fg="#1976d2")
    window.update_idletasks()
    xu_ly_file_excel2(input_file_tab2, output_folder_tab2, ten_file)


def xu_ly_file_excel(excel_file, folder_out, filename):
    try:
        from main import process_excel

        status_var_export_t3.set("⏳ Đang xuất file, vui lòng đợi...")
        status_label_export_t3.config(fg="#1976d2")
        process_excel(
            excel_file,
            folder_out,
            filename,
            allowed_late_evening_staff,
            allowed_late_morning_staff,
        )
        status_var_export_t3.set("✅ Đã xuất file thành công!")
        status_label_export_t3.config(fg="#388e3c")
        file_path = os.path.join(folder_out, filename)
        try:
            os.startfile(file_path)
        except Exception:
            pass
    except Exception as e:
        status_var_export_t3.set(f"❌ Có lỗi xảy ra: {e}")
        status_label_export_t3.config(fg="#b71c1c")


def xu_ly_file_excel2(excel_file, folder_out, filename):
    try:
        from main import process_excel

        status_var_export_t12.set("⏳ Đang xuất file, vui lòng đợi...")
        status_label_export_t12.config(fg="#1976d2")
        process_excel(
            excel_file,
            folder_out,
            filename,
            allowed_late_evening_staff_T12,
            allowed_late_morning_staff_T12,
        )
        status_var_export_t12.set("✅ Đã xuất file thành công!")
        status_label_export_t12.config(fg="#388e3c")
        file_path = os.path.join(folder_out, filename)
        try:
            os.startfile(file_path)
        except Exception:
            pass
    except Exception as e:
        status_var_export_t12.set(f"❌ Có lỗi xảy ra: {e}")
        status_label_export_t12.config(fg="#b71c1c")


# == DEF TAB RULE ===


def tao_tab_quy_dinh(tab):
    tab.grid_rowconfigure(0, weight=1)
    tab.grid_columnconfigure(0, weight=1)

    # Màu sắc và style
    bg_color = "#f6fafd"
    card_colors = [
        ("#e3f2fd", "#1976d2"),  # blue
        ("#fffde7", "#fbc02d"),  # yellow
        ("#ede7f6", "#7e57c2"),  # purple
        ("#e8f5e9", "#388e3c"),  # green
        ("#fff3e0", "#f57c00"),  # orange
    ]
    border_color = "#b3c6e0"
    shadow_color = "#dbe6f6"
    title_fg = "#0d47a1"
    content_fg = "#222"
    font_title = ("Segoe UI", 22, "bold")
    font_card_title = ("Segoe UI", 15, "bold")
    font_content = ("Segoe UI", 12)

    # Canvas để scroll
    canvas = tk.Canvas(tab, borderwidth=0, highlightthickness=0, bg=bg_color)
    frame = tk.Frame(canvas, bg=bg_color)
    scrollbar = ttk.Scrollbar(tab, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=scrollbar.set)

    scrollbar.grid(row=0, column=1, sticky="ns")
    canvas.grid(row=0, column=0, sticky="nsew")
    canvas.create_window((0, 0), window=frame, anchor="nw")

    def on_frame_configure(event):
        canvas.configure(scrollregion=canvas.bbox("all"))

    frame.bind("<Configure>", on_frame_configure)

    # --- Mouse wheel scroll support (Windows, Mac, Linux) ---
    def _on_mousewheel(event):
        if event.delta:
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        elif event.num == 4:
            canvas.yview_scroll(-1, "units")
        elif event.num == 5:
            canvas.yview_scroll(1, "units")

    # Bind mousewheel for Windows and Mac
    canvas.bind_all("<MouseWheel>", _on_mousewheel)
    # Bind mousewheel for Linux
    canvas.bind_all("<Button-4>", _on_mousewheel)
    canvas.bind_all("<Button-5>", _on_mousewheel)

    # Tiêu đề chính cực lớn, có icon (tách icon và chữ để không lỗi font)
    title_frame = tk.Frame(frame, bg=bg_color)
    title_frame.grid(row=0, column=0, sticky="w", padx=40, pady=(24, 18))
    tk.Label(
        title_frame, text="📌", font=("Segoe UI Emoji", 32), bg=bg_color, fg=title_fg
    ).pack(side="left", padx=(0, 12))
    tk.Label(
        title_frame,
        text="QUY ĐỊNH CHẤM CÔNG",
        font=font_title,
        fg=title_fg,
        bg=bg_color,
    ).pack(side="left")

    # Danh sách quy định (card style)
    quy_dinh_sections = [
        {
            "title": "💰 Quy định Sử dụng hệ thống",
            "content": [
                "• Chỉ được chấm công sớm tối đa 15 phút trước khi bắt đầu ca",
                "• Tên nhân viên phải được đặt đúng theo định dạng viết Hoa",
                "• Khi cấu hình ca gãy cần ghi luôn cả ca thường (nếu có lịch của ca thường)",
                "• Nếu tăng ca, cần báo trước cho chủ nhóm để được ghi nhận",
                "• Nếu có lý do đi trễ, cần báo lại để được xử lý hợp lệ",
            ],
        },
        {
            "title": "🕒 Ca sáng (07:00 – 15:00)",
            "content": [
                "• Mỗi phút đi trễ bị trừ 5.000 VNĐ; nếu trễ từ 30 phút thì trừ cố định 200.000 VNĐ",
                "• Nếu đến trước 07:00 thì vẫn tính thời gian làm từ 07:00 (không cộng thêm giờ)",
                "• Ra về sớm không bị trừ lương, nhưng chỉ được tính lương theo số giờ làm thực tế",
            ],
        },
        {
            "title": "🌙 Ca chiều (15:00 – 23:00)",
            "content": [
                "• Áp dụng quy định tương tự như ca sáng",
                "• Nếu đến trước 15:00 thì vẫn tính thời gian làm từ 15:00",
                "• Ra về sớm không bị trừ lương, nhưng chỉ tính theo số giờ làm thực tế",
            ],
        },
        {
            "title": "🔀 Ca gãy linh hoạt",
            "content": [
                "• Nhân viên chỉ được chấm công sớm tối đa 15 phút trước giờ ca gãy đã đăng ký",
                "• Nếu đến đúng giờ linh hoạt đã được phê duyệt: ✅ Được tính công bình thường",
                "• Nếu đến trễ so với giờ linh hoạt: ❌ Bị trừ 5.000 VNĐ/phút; nếu quá 30 phút thì trừ 200.000 VNĐ",
                "• Giờ kết thúc ca là cố định: 15:00 (ca sáng) hoặc 23:00 (ca chiều)",
            ],
        },
        {
            "title": "💰 Cách tính lương",
            "content": [
                "• Lương = Tổng số giờ làm × Đơn giá theo giờ",
                "• Giờ làm chỉ được tính trong khung giờ hợp lệ của từng ca",
                "• File Excel xuất ra sẽ ghi rõ số giờ làm, số tiền lương và các khoản trừ nếu có",
            ],
        },
    ]

    for idx, section in enumerate(quy_dinh_sections):
        card_bg, accent = card_colors[idx % len(card_colors)]
        card_shadow = shadow_color
        # Shadow effect
        card_shadow_frame = tk.Frame(frame, bg=card_shadow)
        card_shadow_frame.grid(
            row=idx + 1, column=0, sticky="ew", padx=38, pady=(0, 24)
        )
        # Card
        card = tk.Frame(
            card_shadow_frame,
            bg=card_bg,
            bd=0,
            highlightbackground=border_color,
            highlightthickness=2,
        )
        card.grid(row=0, column=0, sticky="ewns")
        card.grid_columnconfigure(0, weight=1)
        # Title
        tk.Label(
            card,
            text=section["title"],
            font=font_card_title,
            fg=accent,
            bg=card_bg,
            anchor="w",
        ).grid(row=0, column=0, sticky="w", padx=18, pady=(18, 8))
        # Content
        for i, item in enumerate(section["content"]):
            tk.Label(
                card,
                text=item,
                font=font_content,
                fg=content_fg,
                bg=card_bg,
                anchor="w",
                wraplength=820,
                justify="left",
            ).grid(row=i + 1, column=0, sticky="w", padx=36, pady=2)


# ================== GIAO DIỆN ==================

window = tk.Tk()
window.title("Chấm Công - Tính Lương")
window.geometry("1060x780")
window.minsize(1060, 780)
window_width = 1060
window_height = 780

# Lấy kích thước màn hình
screen_width = window.winfo_screenwidth()
screen_height = window.winfo_screenheight()

# Tính vị trí để căn giữa
x = (screen_width // 2) - (window_width // 2)
y = (screen_height // 2) - (window_height // 2) - 40

# Cập nhật vị trí
window.geometry(f"{window_width}x{window_height}+{x}+{y}")

title_font = ("Segoe UI", 14, "bold")
default_font = ("Segoe UI", 12)
title_font_big = ("Segoe UI", 18, "bold")

# === TAB UI ĐẸP ===
notebook = ttk.Notebook(window)
style = ttk.Style()
style.theme_use("clam")

# Màu sắc cho tab
tab_bg = "#eaf1fb"
tab_selected_bg = "#1976D2"
tab_selected_fg = "#fff"
tab_hover_bg = "#b3c6e0"
tab_border = "#b3c6e0"

# Style cho tab
style.configure("TNotebook", background="#f4f8fb", borderwidth=0)
style.configure(
    "TNotebook.Tab",
    font=("Segoe UI", 12, "bold"),
    padding=[30, 12],
    background=tab_bg,
    foreground="#222",
    borderwidth=0,
    focuscolor=tab_border,
    lightcolor=tab_border,
    bordercolor=tab_border,
    relief="flat",
)
style.map(
    "TNotebook.Tab",
    background=[
        ("selected", tab_selected_bg),
        ("active", tab_hover_bg),
        ("!selected", tab_bg),
    ],
    foreground=[
        ("selected", tab_selected_fg),
        ("active", "#222"),
        ("!selected", "#222"),
    ],
    expand=[("selected", [1, 1, 1, 0])],
)

notebook.pack(fill="both", expand=True, padx=16, pady=12)


tab1 = tk.Frame(notebook, bg="#f5faff")
tab2 = tk.Frame(notebook, bg="#f5faff")
tab3 = tk.Frame(notebook, bg="#f4f8fb")

notebook.add(tab1, text="T3 - Quán Coffee")
notebook.add(tab2, text="T12 - Quán Cơm")
create_tab_gui_mail(notebook)
notebook.add(tab3, text="Quy định")
tao_tab_quy_dinh(tab3)


# ==== TAB T3 ====
tab1.configure(bg="#f7fafd")
label_tab1_title = tk.Label(
    tab1,
    text="T3 - Quán Coffee",
    font=("Segoe UI", 17, "bold"),
    bg="#f7fafd",
    fg="#1a237e",
)
label_tab1_title.grid(row=0, column=0, columnspan=2, pady=(16, 6), sticky="ew")


frame_input = tk.LabelFrame(
    tab1,
    text="📝 Bước 1: Chọn file",
    font=title_font,
    bg="#fafdff",
    fg="#222",
    bd=1,
    relief="groove",
    padx=14,
    pady=10,
    labelanchor="nw",
    highlightthickness=1,
    highlightbackground="#b3c6e0",
)
frame_input.grid(row=1, column=0, columnspan=2, sticky="ew", padx=22, pady=8)
frame_input.columnconfigure(1, weight=1)

btn_chon_file = tk.Button(
    frame_input,
    text="📂 Chọn file Excel",
    font=default_font,
    bg="#f5faff",
    fg="#222",
    activebackground="#e3eafc",
    activeforeground="#222",
    relief="ridge",
    bd=1,
    cursor="hand2",
    command=chon_file_excel,
)
btn_chon_file.grid(row=0, column=0, padx=8, pady=5, ipadx=4, ipady=1)

label_file = tk.Label(
    frame_input, text="Chưa chọn file", font=default_font, bg="#ffffff", fg="#ff1744"
)
label_file.grid(row=0, column=1, sticky="w", padx=8)


# Khung cấu hình ca gãy rõ ràng, border dịu, padding rộng
frame_ca_gay = tk.LabelFrame(
    tab1,
    text="⚙️ Bước 2: Cấu hình Ca Gãy",
    font=title_font,
    bg="#fafdff",
    fg="#1a237e",
    bd=1,
    relief="groove",
    padx=14,
    pady=10,
    labelanchor="nw",
    highlightthickness=1,
    highlightbackground="#b3c6e0",
)
frame_ca_gay.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=22, pady=8)
frame_ca_gay.columnconfigure(0, weight=1)
frame_ca_gay.columnconfigure(1, weight=1)

# === Ca GÃY SÁNG ===

# Ca gãy sáng: khung sáng nhẹ, border mảnh
frame_sang = tk.Frame(
    frame_ca_gay,
    bg="#f7fbff",
    bd=1,
    relief="solid",
    highlightbackground="#b3c6e0",
    highlightthickness=1,
)
frame_sang.grid(row=0, column=0, sticky="nsew", padx=6, pady=4)

tk.Label(
    frame_sang,
    text="Ca Gãy Sáng:",
    font=("Segoe UI", 13, "bold"),
    bg="#f7fbff",
    fg="#1976d2",
).grid(row=0, column=0, sticky="w", padx=4, pady=2)

# Nhập sáng
labels_sang = ["🌅 Nhân viên (VD: Nguyen Van A):", "Giờ bắt đầu (07:00, 08:00,...):"]
for i, label in enumerate(labels_sang):
    tk.Label(frame_sang, text=label, font=default_font, bg="#f7fbff", fg="#222").grid(
        row=i + 1, column=0, sticky="w", padx=4, pady=2
    )

# Combobox chọn tên nhân viên (thay cho Entry)
ten_nv_sang_var = tk.StringVar()
combobox_ten_nv_sang_tab1 = ttk.Combobox(
    frame_sang, textvariable=ten_nv_sang_var, font=default_font, state="readonly"
)
combobox_ten_nv_sang_tab1.grid(row=1, column=1, sticky="ew", padx=4, pady=2)
combobox_ten_nv_sang_tab1["values"] = []

entry_gio_sang_tab1 = tk.Entry(
    frame_sang,
    font=default_font,
    bg="#ffffff",
    relief="solid",
    bd=1,
    highlightthickness=1,
)
entry_gio_sang_tab1.grid(row=2, column=1, sticky="ew", padx=4, pady=2)
entry_gio_sang_tab1.config(highlightcolor="#1976d2", highlightbackground="#b3c6e0")


btn_luu_sang = tk.Button(
    frame_sang,
    text="✅ Lưu Ca Gãy Sáng",
    font=default_font,
    bg="#f5faff",
    fg="#1976d2",
    activebackground="#e3eafc",
    activeforeground="#1976d2",
    relief="ridge",
    bd=1,
    cursor="hand2",
    command=lambda: (
        luu_cau_hinh(
            ten_nv_sang_var.get(),
            entry_gio_sang_tab1.get(),
            allowed_late_morning_staff,
            listbox_gay_sang_tab1,
            "Sáng",
        ),
        luu_du_lieu_ca_gay(),
    ),
)
btn_luu_sang.grid(row=3, column=0, columnspan=2, pady=(5, 2), ipadx=4)


listbox_gay_sang_tab1 = tk.Listbox(
    frame_sang,
    font=default_font,
    height=6,
    bg="#fafdff",
    bd=1,
    relief="solid",
    highlightthickness=1,
    highlightbackground="#e3eaf3",
)
listbox_gay_sang_tab1.grid(row=4, column=0, columnspan=2, sticky="nsew", padx=4, pady=2)


btn_xoa_sang = tk.Button(
    frame_sang,
    text="🗑️ Xóa Ca Gãy Sáng",
    font=default_font,
    bg="#f5faff",
    fg="#b71c1c",
    activebackground="#ffeaea",
    activeforeground="#b71c1c",
    relief="ridge",
    bd=1,
    cursor="hand2",
    command=lambda: xoa_cau_hinh(
        listbox_gay_sang_tab1, allowed_late_morning_staff, "sang", "T3"
    ),
)
btn_xoa_sang.grid(row=5, column=0, columnspan=2, pady=5, ipadx=4)

frame_sang.columnconfigure(1, weight=1)
frame_sang.rowconfigure(3, weight=1)

# === CA CHIỀU GÃY ====

# Ca gãy chiều: khung sáng nhẹ, border mảnh
frame_chieu = tk.Frame(
    frame_ca_gay,
    bg="#fdfcf7",
    bd=1,
    relief="solid",
    highlightbackground="#b3c6e0",
    highlightthickness=1,
)
frame_chieu.grid(row=0, column=1, sticky="nsew", padx=6, pady=4)

tk.Label(
    frame_chieu,
    text="Ca Gãy Chiều:",
    font=("Segoe UI", 13, "bold"),
    bg="#fdfcf7",
    fg="#1976d2",
).grid(row=0, column=0, sticky="w", padx=4, pady=2)


# Nhập chiều
labels_chieu = ["🌇 Nhân viên (VD: Nguyen Van A):", "Giờ bắt đầu (15:00, 16:00,...):"]
for i, label in enumerate(labels_chieu):
    tk.Label(frame_chieu, text=label, font=default_font, bg="#fdfcf7", fg="#222").grid(
        row=i + 1, column=0, sticky="w", padx=4, pady=2
    )

# Combobox chọn tên nhân viên cho ca gãy chiều
ten_nv_chieu_var = tk.StringVar()
combobox_ten_nv_chieu_tab1 = ttk.Combobox(
    frame_chieu, textvariable=ten_nv_chieu_var, font=default_font, state="readonly"
)
combobox_ten_nv_chieu_tab1.grid(row=1, column=1, sticky="ew", padx=4, pady=2)
combobox_ten_nv_chieu_tab1["values"] = []

entry_gio = tk.Entry(
    frame_chieu,
    font=default_font,
    bg="#ffffff",
    relief="solid",
    bd=1,
    highlightthickness=1,
)
entry_gio.grid(row=2, column=1, sticky="ew", padx=4, pady=2)
entry_gio.config(highlightcolor="#fbc02d", highlightbackground="#e0b3b3")

btn_luu_chieu = tk.Button(
    frame_chieu,
    text="✅ Lưu Ca Gãy Chiều",
    font=default_font,
    bg="#f5faff",
    fg="#1976d2",
    activebackground="#e3eafc",
    activeforeground="#1976d2",
    relief="ridge",
    bd=1,
    cursor="hand2",
    command=lambda: (
        luu_cau_hinh(
            ten_nv_chieu_var.get(),
            entry_gio.get(),
            allowed_late_evening_staff,
            listbox_gay,
            "Chiều",
        ),
        luu_du_lieu_ca_gay(),
    ),
)
btn_luu_chieu.grid(row=3, column=0, columnspan=2, pady=(5, 2), ipadx=4)


listbox_gay = tk.Listbox(
    frame_chieu,
    font=default_font,
    height=6,
    bg="#fffdfa",
    bd=1,
    relief="solid",
    highlightthickness=1,
    highlightbackground="#f0e3e3",
)
listbox_gay.grid(row=4, column=0, columnspan=2, sticky="nsew", padx=4, pady=2)


btn_xoa_chieu = tk.Button(
    frame_chieu,
    text="🗑️ Xóa Ca Gãy Chiều",
    font=default_font,
    bg="#f5faff",
    fg="#b71c1c",
    activebackground="#ffeaea",
    activeforeground="#b71c1c",
    relief="ridge",
    bd=1,
    cursor="hand2",
    command=lambda: xoa_cau_hinh(
        listbox_gay, allowed_late_evening_staff, "chieu", "T3"
    ),
)
btn_xoa_chieu.grid(row=5, column=0, columnspan=2, pady=5, ipadx=4)

frame_chieu.columnconfigure(1, weight=1)
frame_chieu.rowconfigure(3, weight=1)

# === KHUNG LƯU FILE ===

# Khung xuất file rõ ràng, border dịu, padding rộng
frame_output = tk.LabelFrame(
    tab1,
    text="💾 Bước 3: Xuất File",
    font=title_font,
    bg="#fafdff",
    fg="#1a237e",
    bd=1,
    relief="groove",
    padx=14,
    pady=10,
    labelanchor="nw",
    highlightthickness=1,
    highlightbackground="#b3c6e0",
)
frame_output.grid(row=3, column=0, columnspan=2, sticky="ew", padx=22, pady=8)
frame_output.columnconfigure(1, weight=1)

# --- Status label for export (T3) ---
status_var_export_t3 = tk.StringVar()
status_label_export_t3 = tk.Label(
    frame_output,
    textvariable=status_var_export_t3,
    font=("Segoe UI", 11),
    fg="#1976d2",
    bg="#fafdff",
)

tk.Label(
    frame_output, text="Tên file (.xlsx):", font=default_font, bg="#ffffff", fg="#222"
).grid(row=0, column=0, sticky="w", padx=5)


# --- Placeholder cho ô nhập tên file (tab T3) ---
def set_placeholder_t3(event=None):
    if not entry_ten_file.get():
        entry_ten_file.insert(0, "Nhập tên file xuất ra (.xlsx)")
        entry_ten_file.config(fg="#888")


def clear_placeholder_t3(event=None):
    if entry_ten_file.get() == "Nhập tên file xuất ra (.xlsx)":
        entry_ten_file.delete(0, tk.END)
        entry_ten_file.config(fg="#222")


entry_ten_file = tk.Entry(
    frame_output,
    font=default_font,
    width=40,
    state="normal",
    bg="#ffffff",
    relief="solid",
    bd=1,
    highlightthickness=1,
    highlightbackground="#b3c6e0",
)
entry_ten_file.grid(row=0, column=1, padx=5, sticky="ew")
set_placeholder_t3()
entry_ten_file.bind("<FocusIn>", clear_placeholder_t3)
entry_ten_file.bind("<FocusOut>", set_placeholder_t3)


btn_chon_thu_muc = tk.Button(
    frame_output,
    text="💾 Chọn nơi lưu file",
    font=default_font,
    bg="#f5faff",
    fg="#1a237e",
    activebackground="#e3eafc",
    activeforeground="#1a237e",
    state="disabled",
    relief="ridge",
    bd=1,
    cursor="hand2",
    command=chon_thu_muc,
)
btn_chon_thu_muc.grid(row=1, column=0, pady=10, ipadx=4, sticky="w")
status_label_export_t3.grid(row=1, column=1, sticky="w", padx=(10, 0))


label_folder = tk.Label(
    frame_output,
    text="Chưa chọn nơi lưu",
    font=default_font,
    bg="#ffffff",
    fg="#b71c1c",
)
label_folder.grid(row=2, column=0, columnspan=2)


# ==== CẢI TIẾN UI ĐƠN GIẢN, RÕ RÀNG CHO TAB T12 ====
tab2.configure(bg="#f7fafd")
label_tab_title2 = tk.Label(
    tab2,
    text="T12 - Quán Cơm",
    font=("Segoe UI", 17, "bold"),
    bg="#f7fafd",
    fg="#1a237e",
)
label_tab_title2.grid(row=0, column=0, columnspan=2, pady=(16, 6), sticky="ew")


# Khung chọn file rõ ràng, bo góc nhẹ, border dịu
frame_input2 = tk.LabelFrame(
    tab2,
    text="📝 Bước 1: Chọn file",
    font=title_font,
    bg="#fafdff",
    fg="#222",
    bd=1,
    relief="groove",
    padx=14,
    pady=10,
    labelanchor="nw",
    highlightthickness=1,
    highlightbackground="#b3c6e0",
)
frame_input2.grid(row=1, column=0, columnspan=2, sticky="ew", padx=22, pady=8)
frame_input2.columnconfigure(1, weight=1)

btn_chon_file2 = tk.Button(
    frame_input2,
    text="📂 Chọn file Excel",
    font=default_font,
    bg="#f5faff",
    fg="#222",
    activebackground="#e3eafc",
    activeforeground="#222",
    relief="ridge",
    bd=1,
    cursor="hand2",
    command=chon_file_excel_T12,
)
btn_chon_file2.grid(row=0, column=0, padx=8, pady=5, ipadx=4, ipady=1)

label_file2 = tk.Label(
    frame_input2, text="Chưa chọn file", font=default_font, bg="#ffffff", fg="#ff1744"
)
label_file2.grid(row=0, column=1, sticky="w", padx=8)


# Khung cấu hình ca gãy rõ ràng, border dịu, padding rộng
frame_ca_gay2 = tk.LabelFrame(
    tab2,
    text="⚙️ Bước 2: Cấu hình Ca Gãy",
    font=title_font,
    bg="#fafdff",
    fg="#1a237e",
    bd=1,
    relief="groove",
    padx=14,
    pady=10,
    labelanchor="nw",
    highlightthickness=1,
    highlightbackground="#b3c6e0",
)
frame_ca_gay2.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=22, pady=8)
frame_ca_gay2.columnconfigure(0, weight=1)
frame_ca_gay2.columnconfigure(1, weight=1)

# === Ca GÃY SÁNG TAB 2 ===

# Ca gãy sáng: khung sáng nhẹ, border mảnh
frame_sang2 = tk.Frame(
    frame_ca_gay2,
    bg="#f7fbff",
    bd=1,
    relief="solid",
    highlightbackground="#b3c6e0",
    highlightthickness=1,
)
frame_sang2.grid(row=0, column=0, sticky="nsew", padx=6, pady=4)

tk.Label(
    frame_sang2,
    text="Ca Gãy Sáng:",
    font=("Segoe UI", 13, "bold"),
    bg="#f7fbff",
    fg="#1976d2",
).grid(row=0, column=0, sticky="w", padx=4, pady=2)

tk.Label(
    frame_sang2,
    text="🌅 Nhân viên (VD: Nguyen Van A):",
    font=default_font,
    bg="#f7fbff",
    fg="#222",
).grid(row=1, column=0, sticky="w", padx=4, pady=2)
# Combobox chọn tên nhân viên cho ca gãy sáng tab T12
ten_nv_sang_tab2_var = tk.StringVar()
combobox_ten_nv_sang_tab2 = ttk.Combobox(
    frame_sang2, textvariable=ten_nv_sang_tab2_var, font=default_font, state="readonly"
)
combobox_ten_nv_sang_tab2.grid(row=1, column=1, sticky="ew", padx=4, pady=2)
combobox_ten_nv_sang_tab2["values"] = []

tk.Label(
    frame_sang2,
    text="Giờ bắt đầu (07:00, 08:00,...):",
    font=default_font,
    bg="#f7fbff",
    fg="#222",
).grid(row=2, column=0, sticky="w", padx=4, pady=2)
entry_gio_sang_tab2 = tk.Entry(
    frame_sang2,
    font=default_font,
    bg="#ffffff",
    relief="solid",
    bd=1,
    highlightthickness=1,
    highlightbackground="#b3c6e0",
)
entry_gio_sang_tab2.grid(row=2, column=1, sticky="ew", padx=4, pady=2)

btn_luu_sang_tab2 = tk.Button(
    frame_sang2,
    text="✅ Lưu Ca Gãy Sáng",
    font=default_font,
    bg="#f5faff",
    fg="#1976d2",
    activebackground="#e3eafc",
    activeforeground="#1976d2",
    relief="ridge",
    bd=1,
    cursor="hand2",
    command=lambda: (
        luu_cau_hinh(
            ten_nv_sang_tab2_var.get(),
            entry_gio_sang_tab2.get(),
            allowed_late_morning_staff_T12,
            listbox_gay_sang_tab2,
            "Sáng",
        ),
        luu_du_lieu_ca_gay_T12(),
    ),
)
btn_luu_sang_tab2.grid(row=3, column=0, columnspan=2, pady=(5, 2), ipadx=4)


listbox_gay_sang_tab2 = tk.Listbox(
    frame_sang2,
    font=default_font,
    height=6,
    bg="#fafdff",
    bd=1,
    relief="solid",
    highlightthickness=1,
    highlightbackground="#e3eaf3",
)
listbox_gay_sang_tab2.grid(row=4, column=0, columnspan=2, sticky="nsew", padx=4, pady=2)


btn_xoa_sang_tab2 = tk.Button(
    frame_sang2,
    text="🗑️ Xóa Ca Gãy Sáng",
    font=default_font,
    bg="#f5faff",
    fg="#b71c1c",
    activebackground="#ffeaea",
    activeforeground="#b71c1c",
    relief="ridge",
    bd=1,
    cursor="hand2",
    command=lambda: xoa_cau_hinh(
        listbox_gay_sang_tab2, allowed_late_morning_staff_T12, "sang", "T12"
    ),
)
btn_xoa_sang_tab2.grid(row=5, column=0, columnspan=2, pady=5, ipadx=4)

frame_sang2.columnconfigure(1, weight=1)

# === Ca GÃY CHIỀU TAB 2 ===

# Ca gãy chiều: khung sáng nhẹ, border mảnh
frame_chieu2 = tk.Frame(
    frame_ca_gay2,
    bg="#fdfcf7",
    bd=1,
    relief="solid",
    highlightbackground="#b3c6e0",
    highlightthickness=1,
)
frame_chieu2.grid(row=0, column=1, sticky="nsew", padx=6, pady=4)

tk.Label(
    frame_chieu2,
    text="Ca Gãy Chiều:",
    font=("Segoe UI", 13, "bold"),
    bg="#fdfcf7",
    fg="#1976d2",
).grid(row=0, column=0, sticky="w", padx=4, pady=2)

tk.Label(
    frame_chieu2,
    text="🌇 Nhân viên (VD: Nguyen Van A):",
    font=default_font,
    bg="#fdfcf7",
    fg="#222",
).grid(row=1, column=0, sticky="w", padx=4, pady=2)
# Combobox chọn tên nhân viên cho ca gãy chiều tab T12
ten_nv_chieu_tab2_var = tk.StringVar()
combobox_ten_nv_chieu_tab2 = ttk.Combobox(
    frame_chieu2,
    textvariable=ten_nv_chieu_tab2_var,
    font=default_font,
    state="readonly",
)
combobox_ten_nv_chieu_tab2.grid(row=1, column=1, sticky="ew", padx=4, pady=2)
combobox_ten_nv_chieu_tab2["values"] = []

tk.Label(
    frame_chieu2,
    text="Giờ bắt đầu (15:00, 16:00,...):",
    font=default_font,
    bg="#fdfcf7",
    fg="#222",
).grid(row=2, column=0, sticky="w", padx=4, pady=2)
entry_gio_chieu_tab2 = tk.Entry(
    frame_chieu2,
    font=default_font,
    bg="#ffffff",
    relief="solid",
    bd=1,
    highlightthickness=1,
    highlightbackground="#e0b3b3",
)
entry_gio_chieu_tab2.grid(row=2, column=1, sticky="ew", padx=4, pady=2)


btn_luu_chieu_tab2 = tk.Button(
    frame_chieu2,
    text="✅ Lưu Ca Gãy Chiều",
    font=default_font,
    bg="#f5faff",
    fg="#fbc02d",
    activebackground="#fffde7",
    activeforeground="#fbc02d",
    relief="ridge",
    bd=1,
    cursor="hand2",
    command=lambda: (
        luu_cau_hinh(
            ten_nv_chieu_tab2_var.get(),
            entry_gio_chieu_tab2.get(),
            allowed_late_evening_staff_T12,
            listbox_gay_chieu_tab2,
            "Chiều",
        ),
        luu_du_lieu_ca_gay_T12(),
    ),
)
btn_luu_chieu_tab2.config(
    fg="#1976d2", activebackground="#e3eafc", activeforeground="#1976d2"
)
btn_luu_chieu_tab2.grid(row=3, column=0, columnspan=2, pady=(5, 2), ipadx=4)


listbox_gay_chieu_tab2 = tk.Listbox(
    frame_chieu2,
    font=default_font,
    height=6,
    bg="#fffdfa",
    bd=1,
    relief="solid",
    highlightthickness=1,
    highlightbackground="#f0e3e3",
)
listbox_gay_chieu_tab2.grid(
    row=4, column=0, columnspan=2, sticky="nsew", padx=4, pady=2
)


btn_xoa_chieu_tab2 = tk.Button(
    frame_chieu2,
    text="🗑️ Xóa Ca Gãy Chiều",
    font=default_font,
    bg="#f5faff",
    fg="#b71c1c",
    activebackground="#ffeaea",
    activeforeground="#b71c1c",
    relief="ridge",
    bd=1,
    cursor="hand2",
    command=lambda: xoa_cau_hinh(
        listbox_gay_chieu_tab2, allowed_late_evening_staff_T12, "chieu", "T12"
    ),
)
btn_xoa_chieu_tab2.grid(row=5, column=0, columnspan=2, pady=5, ipadx=4)

frame_chieu2.columnconfigure(1, weight=1)

# === KHUNG XUẤT FILE TAB 2 ===

# Khung xuất file rõ ràng, border dịu, padding rộng
frame_output2 = tk.LabelFrame(
    tab2,
    text="💾 Bước 3: Xuất File",
    font=title_font,
    bg="#fafdff",
    fg="#1a237e",
    bd=1,
    relief="groove",
    padx=14,
    pady=10,
    labelanchor="nw",
    highlightthickness=1,
    highlightbackground="#b3c6e0",
)
frame_output2.grid(row=3, column=0, columnspan=2, sticky="ew", padx=22, pady=8)
frame_output2.columnconfigure(1, weight=1)

# --- Status label for export (T12) ---
status_var_export_t12 = tk.StringVar()
status_label_export_t12 = tk.Label(
    frame_output2,
    textvariable=status_var_export_t12,
    font=("Segoe UI", 11),
    fg="#1976d2",
    bg="#fafdff",
)


tk.Label(
    frame_output2, text="Tên file (.xlsx):", font=default_font, bg="#ffffff", fg="#222"
).grid(row=0, column=0, sticky="w", padx=5)


# --- Placeholder cho ô nhập tên file (tab T12) ---
def set_placeholder_t12(event=None):
    if not entry_ten_file2.get():
        entry_ten_file2.insert(0, "Nhập tên file xuất ra (.xlsx)")
        entry_ten_file2.config(fg="#888")


def clear_placeholder_t12(event=None):
    if entry_ten_file2.get() == "Nhập tên file xuất ra (.xlsx)":
        entry_ten_file2.delete(0, tk.END)
        entry_ten_file2.config(fg="#222")


entry_ten_file2 = tk.Entry(
    frame_output2,
    font=default_font,
    width=40,
    state="normal",
    bg="#ffffff",
    relief="solid",
    bd=1,
    highlightthickness=1,
    highlightbackground="#b3c6e0",
)
entry_ten_file2.grid(row=0, column=1, padx=5, sticky="ew")
set_placeholder_t12()
entry_ten_file2.bind("<FocusIn>", clear_placeholder_t12)
entry_ten_file2.bind("<FocusOut>", set_placeholder_t12)


btn_chon_thu_muc2 = tk.Button(
    frame_output2,
    text="💾 Chọn nơi lưu file",
    font=default_font,
    bg="#f5faff",
    fg="#1a237e",
    activebackground="#e3eafc",
    activeforeground="#1a237e",
    state="disabled",
    relief="ridge",
    bd=1,
    cursor="hand2",
    command=chon_thu_muc_tab2,
)
btn_chon_thu_muc2.grid(row=1, column=0, pady=10, ipadx=4, sticky="w")
status_label_export_t12.grid(row=1, column=1, sticky="w", padx=(10, 0))


label_folder2 = tk.Label(
    frame_output2,
    text="Chưa chọn nơi lưu",
    font=default_font,
    bg="#ffffff",
    fg="#b71c1c",
)
label_folder2.grid(row=2, column=0, columnspan=2)


nap_du_lieu_ca_gay()
cap_nhat_danh_sach(listbox_gay_sang_tab1, allowed_late_morning_staff)
cap_nhat_danh_sach(listbox_gay, allowed_late_evening_staff)

nap_du_lieu_ca_gay_T12()
cap_nhat_danh_sach(listbox_gay_sang_tab2, allowed_late_morning_staff_T12)
cap_nhat_danh_sach(listbox_gay_chieu_tab2, allowed_late_evening_staff_T12)
window.mainloop()
